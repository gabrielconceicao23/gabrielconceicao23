#!/usr/bin/env python3
"""Extrai texto de imagens e salva em uma planilha XLSX."""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Iterable

import pytesseract
from PIL import Image
from openpyxl import Workbook


EXTENSOES_SUPORTADAS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
PADRAO_NUMERO = r"\d{1,3}(?:\.\d{3})*,\d+"
PADRAO_FIM_LINHA = re.compile(
    rf"(?P<preco_normal>{PADRAO_NUMERO})\s+"
    rf"(?P<quantidade>{PADRAO_NUMERO})\s+"
    rf"(?P<valor_total>{PADRAO_NUMERO})$"
)


def listar_imagens(entrada: Path) -> list[Path]:
    """Retorna uma lista ordenada de imagens a partir de arquivo ou pasta."""
    if entrada.is_file():
        if entrada.suffix.lower() not in EXTENSOES_SUPORTADAS:
            raise ValueError(f"Arquivo não suportado: {entrada}")
        return [entrada]

    if not entrada.is_dir():
        raise FileNotFoundError(f"Caminho não encontrado: {entrada}")

    imagens = sorted(
        arquivo
        for arquivo in entrada.iterdir()
        if arquivo.is_file() and arquivo.suffix.lower() in EXTENSOES_SUPORTADAS
    )

    if not imagens:
        raise ValueError(f"Nenhuma imagem suportada encontrada em: {entrada}")

    return imagens


def extrair_texto(caminho_imagem: Path, idioma: str) -> str:
    """Executa OCR usando o Tesseract e devolve o texto extraído."""
    with Image.open(caminho_imagem) as img:
        return pytesseract.image_to_string(img, lang=idioma).strip()


def linhas_sem_vazio(texto: str) -> Iterable[str]:
    for linha in texto.splitlines():
        conteudo = linha.strip()
        if conteudo:
            yield conteudo


def extrair_descricao_unidade(bloco: str) -> tuple[str, str]:
    """
    Separa descrição e unidade no trecho à esquerda dos valores.

    Exemplo esperado:
    - "AFIADOR FACA MOZCADA UN CX C/ 12"
    - "POTE ... 520ML CX C/ 12"
    """
    match = re.search(r"\s+(CX\s+C/\s*\d+)\s*$", bloco)
    if match:
        unidade = match.group(1)
        descricao = bloco[: match.start()].strip()
        return descricao, unidade

    partes = bloco.rsplit(maxsplit=1)
    if len(partes) == 2:
        return partes[0].strip(), partes[1].strip()
    return bloco.strip(), ""


def parse_linha_tabela(linha: str) -> dict[str, str] | None:
    """Tenta converter uma linha OCR em colunas de tabela."""
    if not linha or linha.lower().startswith("produto"):
        return None

    fim = PADRAO_FIM_LINHA.search(linha)
    if not fim:
        return None

    esquerda = linha[: fim.start()].strip()
    match_produto = re.match(r"^(?P<produto>\d{5,})\s+(?P<resto>.+)$", esquerda)
    if not match_produto:
        return None

    produto = match_produto.group("produto")
    resto = match_produto.group("resto").strip()
    descricao, unidade = extrair_descricao_unidade(resto)

    return {
        "produto": produto,
        "descricao": descricao,
        "unidade": unidade,
        "preco_normal": fim.group("preco_normal"),
        "quantidade": fim.group("quantidade"),
        "valor_total": fim.group("valor_total"),
    }


def gerar_xlsx(imagens: list[Path], destino: Path, idioma: str) -> None:
    """Cria planilha com cada linha reconhecida e uma aba de texto completo."""
    wb = Workbook()

    ws_linhas = wb.active
    ws_linhas.title = "linhas"
    ws_linhas.append(["imagem", "n_linha", "texto"])

    ws_texto = wb.create_sheet(title="texto_completo")
    ws_texto.append(["imagem", "texto"])

    ws_tabela = wb.create_sheet(title="tabela")
    ws_tabela.append(
        [
            "imagem",
            "produto",
            "descricao",
            "unidade",
            "preco_normal",
            "quantidade",
            "valor_total",
        ]
    )

    for imagem in imagens:
        texto = extrair_texto(imagem, idioma=idioma)
        ws_texto.append([imagem.name, texto])

        for indice, linha in enumerate(linhas_sem_vazio(texto), start=1):
            ws_linhas.append([imagem.name, indice, linha])
            parsed = parse_linha_tabela(linha)
            if parsed:
                ws_tabela.append(
                    [
                        imagem.name,
                        parsed["produto"],
                        parsed["descricao"],
                        parsed["unidade"],
                        parsed["preco_normal"],
                        parsed["quantidade"],
                        parsed["valor_total"],
                    ]
                )

    ws_linhas.column_dimensions["A"].width = 30
    ws_linhas.column_dimensions["B"].width = 10
    ws_linhas.column_dimensions["C"].width = 80
    ws_texto.column_dimensions["A"].width = 30
    ws_texto.column_dimensions["B"].width = 100
    ws_tabela.column_dimensions["A"].width = 30
    ws_tabela.column_dimensions["B"].width = 12
    ws_tabela.column_dimensions["C"].width = 60
    ws_tabela.column_dimensions["D"].width = 14
    ws_tabela.column_dimensions["E"].width = 14
    ws_tabela.column_dimensions["F"].width = 14
    ws_tabela.column_dimensions["G"].width = 14

    wb.save(destino)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extrai texto de uma imagem (ou pasta de imagens) e salva em XLSX."
    )
    parser.add_argument(
        "entrada",
        type=Path,
        help="Caminho para imagem ou pasta com imagens.",
    )
    parser.add_argument(
        "saida",
        type=Path,
        nargs="?",
        default=Path("resultado_ocr.xlsx"),
        help="Arquivo XLSX de saída (padrão: resultado_ocr.xlsx).",
    )
    parser.add_argument(
        "--idioma",
        default="por",
        help="Idioma do OCR no Tesseract (ex.: por, eng, spa). Padrão: por.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    imagens = listar_imagens(args.entrada)
    gerar_xlsx(imagens, destino=args.saida, idioma=args.idioma)
    print(f"Planilha criada com sucesso: {args.saida.resolve()}")


if __name__ == "__main__":
    main()
